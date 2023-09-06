"""

 EDITEUR DE MAP CREANT UNE MAP A PARTIR D'UN FICHIER EXCEL

"""

import openpyxl as op

"""
Différentes couleurs avec leur traduction hexadécimal et leur fonctionnalité:
    -vert : emetteur FF00B050
    -bleu : milieu mur FF002060
    -rouge : extremité mur FFFF0000
"""

def editMapBis(name):
    """
    :param name: nom du fichier excel
    :return: renvoie la liste des murs avec leurs caractéristiques et le ou les émetteur(s) avec leur caractéristiques
    """
    posMurH = list()
    posMurV = list()
    posEmetteurs = list()

    wb = op.load_workbook(name, read_only=True)
    ws = wb.active

    for max_row, row in enumerate(ws, 1):
        if all(c.value is None for c in row):
            break

    maxRow = ws.max_row

    # Création des murs horizontaux
    for i in range(ws.min_row, ws.max_row+1):
        value = 0
        listExtremite = list()
        for j in range(ws.min_column, ws.max_column+1):
            try:
                if(ws.cell(i, j).fill.start_color.index == "FFFF0000"):
                    if(len(listExtremite)==4):
                        a = listExtremite[0]
                        b = listExtremite[1]
                        listExtremite = [a,b]
                    if (j != 1):
                        if (not ws.cell(i, j-1).fill.start_color.index == "FFFF0000"):
                            listExtremite.extend([str((j - 1)*dim), str((maxRow - i)*dim)])
                    else:
                        listExtremite.extend([str((j - 1)*dim), str((maxRow - i)*dim)])

                elif(ws.cell(i, j).fill.start_color.index == "FF002060"):

                    val = ws.cell(i, j).value
                    if(val != None):
                        value = val
                else:

                    if(len(listExtremite)==4):
                        width = value.split("/")[0]
                        type = value.split("/")[1]
                        typeMur = checkTypeMur(type)
                        listExtremite.extend([width, typeMur[0], typeMur[1]])
                        posMurH.append(listExtremite)
                    listExtremite = list()

                if(j == ws.max_column):
                    if (len(listExtremite) == 4):
                        width = value.split("/")[0]
                        type = value.split("/")[1]
                        typeMur = checkTypeMur(type)
                        listExtremite.extend([width, typeMur[0], typeMur[1]])
                        posMurH.append(listExtremite[:])
                    listExtremite = list()

                if(ws.cell(i, j).fill.start_color.index == "FF00B050"):
                    posEmetteurs.append([str((j - 1)*dim), str((maxRow - i)*dim), ws.cell(i, j).value])
            except:
                pass

    # Création des murs verticaux
    for i in range(ws.min_column, ws.max_column+1):
        value = 0
        listExtremite = list()
        for j in range(ws.min_row, ws.max_row+1):
            try:
                if(ws.cell(j, i).fill.start_color.index == "FFFF0000"):

                    if (len(listExtremite) == 4):
                        a = listExtremite[0]
                        b = listExtremite[1]
                        listExtremite = [a, b]

                    if(j!=1):
                        if(not ws.cell(j-1, i).fill.start_color.index == "FFFF0000"):
                            listExtremite.extend([str((i - 1)*dim), str((maxRow - j)*dim)])
                    else:
                        listExtremite.extend([str((i - 1)*dim), str((maxRow - j)*dim)])



                elif(ws.cell(j, i).fill.start_color.index == "FF002060"):

                    val = ws.cell(j, i).value
                    if(val != None):
                        value = val


                else:
                    if(len(listExtremite)==4):
                        width = value.split("/")[0]
                        type = value.split("/")[1]
                        typeMur = checkTypeMur(type)
                        listExtremite.extend([width, typeMur[0], typeMur[1]])
                        posMurV.append(listExtremite)
                    listExtremite = list()


                if(j == ws.max_row):
                    if (len(listExtremite) == 4):
                        width = value.split("/")[0]
                        type = value.split("/")[1]
                        typeMur = checkTypeMur(type)
                        listExtremite.extend([width, typeMur[0], typeMur[1]])
                        posMurV.append(listExtremite[:])
                    listExtremite = list()
            except:
                pass
    wb.close()
    posMurH.extend(posMurV)
    createFileMurs(posMurH)
    createFileMap(ws.max_column, ws.max_row, posEmetteurs)

def checkTypeMur(nbre):
    """
    :param nbre: le numéro qui correspond à un matériau précis
    :return: renvoie les caractéristiques du matériau (brique, béton, ou cloison)
    """
    nbre = int(nbre)
    caract = list()
    if(nbre==1):
        #Brique
        caract = ['4.6','0.02']
    elif(nbre==2):
        #Béton
        caract = ['5', '0.014']
    else:
        #Cloison
        caract = ['2.25', '0.04']

    return caract

def createFileMurs(listeMurs):
    """
    :param listeMurs: la liste des murs de la map
    :return: pas de return. On créé le fichier mur
    """
    wall_fichier = open("walls.txt", "w")
    for wall in listeMurs:
        for elem in wall:
            wall_fichier.write(elem)
            wall_fichier.write(' ')
        wall_fichier.write('\n')
    wall_fichier.close()

def createFileMap(maxX, maxY, posEmetteurs):
    """
    :param maxX: la taille en X de la map
    :param maxY: la taille en Y de la map
    :param posEmetteurs: la liste des émetteurs contenant leur position et leur phase
    :return: pas de return. On créé le fichier map
    """
    map_fichier = open("map.txt", "w")
    for posEmet in posEmetteurs:
        map_fichier.write(posEmet[0] + " " + posEmet[1]+ " " + str(posEmet[2]) + "/")
    map_fichier.write("\n"+str((maxX-1)*dim))
    map_fichier.write(" ")
    map_fichier.write(str((maxY-1)*dim))
    map_fichier.write(" ")
    map_fichier.write("100") # Discrétisation
    map_fichier.close()

# Dimension d'une cellule. On multiplie toutes les cellules par la dimension
dim = 4

